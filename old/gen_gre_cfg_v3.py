import openpyxl
from openpyxl import load_workbook
import argparse

# Парсинг аргументов командной строки
parser = argparse.ArgumentParser(description="Generate GRE tunnel configuration for Cisco ASR1002 routers.")
parser.add_argument("router", choices=["r1", "r2"], help="Choose the router (r1 or r2) for which to generate configuration.")
args = parser.parse_args()

# Путь к файлу Excel с данными
excel_file_path = '/mnt/c/Users/SBakhvalov/Documents/wsl_ws/cppk/SB_ip_plan.xlsx'

# Открываем файл Excel с аргументом data_only=True
workbook = load_workbook(excel_file_path, data_only=True)
sheet = workbook.active

# Определяем шаблон конфигурации GRE туннеля для Cisco ASR1002
gre_template = """
interface Tunnel{tunnel_number}
 description {description}
 ip address {source_ip} {subnet_mask}
 tunnel source {source_ip}
 tunnel destination {destination_ip}
"""

# Функция для генерации конфигурации GRE туннелей
def generate_gre_config(router):
    # Проходимся по строкам файла Excel, начиная с второй строки (первая строка - заголовки)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        remote_site = row[0]  # Название удаленной площадки
        source_ip = row[1 if router == "r1" else 4]  # IP адрес туннельного интерфейса для r1 или r2
        subnet_mask = row[2 if router == "r1" else 5]  # Маска подсети для r1 или r2
        destination_ip = row[3 if router == "r1" else 6]  # IP адрес назначения для r1 или r2
        description = f"GRE tunnel to {remote_site}"

        # Генерируем конфигурацию для выбранного роутера
        gre_config = gre_template.format(
            tunnel_number=row[0],  # Номер туннеля соответствует названию удаленной площадки
            description=description,
            source_ip=source_ip,
            subnet_mask=subnet_mask,
            destination_ip=destination_ip
        )

        # Выводим сгенерированную конфигурацию в терминал
        print(gre_config)

# Вызываем функцию для генерации конфигурации для выбранного роутера
generate_gre_config(args.router)
