import openpyxl
from openpyxl import load_workbook

# Путь к файлу Excel с данными
excel_file_path = '/mnt/c/Users/SBakhvalov/Documents/wsl_ws/cppk/SB_ip_plan.xlsx'

# Открываем файл Excel
workbook = load_workbook(excel_file_path)
sheet = workbook.active

# Определяем шаблон конфигурации GRE туннеля для Cisco ASR1002
gre_template = """
interface Tunnel{tunnel_number}
 description {description}
 ip address {source_ip} {subnet_mask}
 tunnel source {source_ip}
 tunnel destination {destination_ip}
"""

# Функция для генерации конфигурации GRE туннелей
def generate_gre_config():
    # Проходимся по строкам файла Excel, начиная с второй строки (первая строка - заголовки)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        remote_site = row[0]  # Название удаленной площадки
        source_ip_1 = row[1]  # IP адрес туннельного интерфейса для RT01-BRD
        subnet_mask_1 = row[2]  # Маска подсети для туннельного интерфейса RT01-BRD
        destination_ip_1 = row[3]  # IP адрес назначения для RT01-BRD
        source_ip_2 = row[4]  # IP адрес туннельного интерфейса для RT02-BRD
        subnet_mask_2 = row[5]  # Маска подсети для туннельного интерфейса RT02-BRD
        destination_ip_2 = row[6]  # IP адрес назначения для RT02-BRD
        description = f"GRE tunnel to {remote_site}"

        # Генерируем конфигурацию для RT01-BRD
        rt01_gre_config = gre_template.format(
            tunnel_number=row[0],  # Номер туннеля соответствует названию удаленной площадки
            description=description,
            source_ip=source_ip_1,
            subnet_mask=subnet_mask_1,
            destination_ip=destination_ip_1
        )

        # Генерируем конфигурацию для RT02-BRD
        rt02_gre_config = gre_template.format(
            tunnel_number=row[0],  # Номер туннеля соответствует названию удаленной площадки
            description=description,
            source_ip=source_ip_2,
            subnet_mask=subnet_mask_2,
            destination_ip=destination_ip_2
        )

        # Здесь можно сохранить конфигурацию в файл или выводить на экран
        # Например, можно добавить код для записи в файл
        # with open('gre_cfg.txt', 'a') as cfg_file:
        #     cfg_file.write(rt01_gre_config)
        #     cfg_file.write(rt02_gre_config)
        print(rt01_gre_config)

# Вызываем функцию для генерации конфигурации
generate_gre_config()
