import openpyxl
import os

# Создаем новую книгу для результатов
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Указываем путь к папке с файлами Excel
folder_path = '/mnt/c/Users/SBakhvalov/Documents/wsl_ws/cppk/polygon'  # Замените на фактический путь к вашей папке
#folder_path = ''  # Замените на фактический путь к вашей папке

# Получаем список файлов Excel в указанной папке
excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
print(excel_files)

# Итерируемся по каждому файлу
for excel_file in excel_files:
        # Полный путь к текущему файлу
    full_path = os.path.join(folder_path, excel_file)
    print(full_path)
    # Открываем текущий файл Excel
    current_workbook = openpyxl.load_workbook(full_path)
    current_sheet = current_workbook.active

    # Ищем ячейку с текстом "Адресация GRE Tun:"
    target_text = "Адресация GRE Tun:"
    found = False

    for row in current_sheet.iter_rows():
        for cell in row:
            if cell.value == target_text:
                found = True
                target_row = cell.row
                target_column = cell.column

    # Если ячейка найдена, извлекаем данные из следующего столбца
    if found:
        for row in current_sheet.iter_rows(min_row=target_row, max_row=current_sheet.max_row, min_col=target_column + 1, max_col=target_column + 1):
            for cell in row:
                # Добавляем название файла без расширения и значение в результаты
                output_sheet.append([os.path.splitext(excel_file)[0], cell.value])

    # Закрываем текущий файл Excel
    current_workbook.close()

# Сохраняем результаты в новом файле
output_workbook.save('output.xlsx')
