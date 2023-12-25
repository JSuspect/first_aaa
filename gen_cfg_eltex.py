import openpyxl
import pathlib

# Пути к рабочей папке и файлу Excel
excel_path = pathlib.Path.home() / 'slink' / 'wsld' / 'rusatom' / 'table.xlsx'
out_dir = pathlib.Path.home() / 'slink' / 'wsld' / 'rusatom' / 'config' / 'pygen_common_conf'
new_excel_path = pathlib.Path.home() / 'slink' / 'wsld' / 'rusatom' / 'modified_table.xlsx'


# Загрузка рабочей книги Excel и листа KJ
excel_book = openpyxl.load_workbook(excel_path, data_only=True)
sheet_main = excel_book['KJ']
sheet_uplinks = excel_book['uplinks']

def gen_service_conf(intf_name, vlan_id):
    return f'\nmac address-table static ec:8e:b5:4b:98:8f vlan {vlan_id} interface {intf_name} secure\n'

# Функция для генерации конфигурации интерфейсов
def generate_interface_config(interface_name, description, mode, vlan, role):
    # Проверяем, задано ли описание

    description_part = f" description {description}\n" if description else ''

    # Подготавливаем общие части шаблона
    common_part = (f'interface {interface_name}\n'
                    f'{description_part}'
                    f' switchport mode {mode}\n'
                    f' switchport access vlan {vlan}\n'
                    f' spanning-tree portfast\n'
                    f' storm-control broadcast level 20 trap shutdown\n'
                    f' storm-control multicast level 20 trap shutdown\n'
                    f' no lldp transmit\n'
                    )
    
    # Подготавливаем часть шаблона для сервисных портов
    service_part = (' port security mode secure permanent\n'
                    ' port security max 1\n'
                    ' port security routed secure-address ec:8e:b5:4b:98:8f\n'
                    ' port security discard\n'
                    ' exit\n'
                    )
    if role == 'service':
        res = gen_service_conf(interface_name, vlan)
        return common_part + service_part + '!\n' + res
    else:
        return common_part + '!\n'

# Функция для сохранения файла конфигурации
def save_configuration(out_dir, current_hostname, current_ip, config_lines, vlan_set, uplink1, uplink2):
    config_interface = '\n'.join(config_lines)
    trunk_vlans = ','.join(str(vlan) for vlan in sorted(vlan_set))
    file_path = out_dir / f'{current_ip}_{current_hostname}.txt'
    conf = config_tpl.format(host = current_hostname, conf_int=config_interface, vlans=trunk_vlans, ip_add=current_ip, u1 = uplink1, u2 = uplink2)

    with file_path.open(mode='w', encoding='utf-8') as config_file:
        config_file.write(conf)

# Полный шаблон конфигурации
config_tpl = '''
vlan 2021 name MGMT_SCADA_ANE
!
vlan database
 vlan 10,12,13,15,2021
exit
!
errdisable recovery interval 120
errdisable recovery cause all
!
hostname {host}
!
clock timezone MSK +3
!
logging buffered warnings
logging cli-commands
logging file informational
!
line console
 exec-timeout 20
exit
!
line telnet
 exec-timeout 20
exit
!
line ssh
 exec-timeout 20
exit
!
line telnet
 login authentication RUSATOM
 enable authentication RUSATOM
 password da39a3ee5e6b4b0d3255bfef95601890afd80709 encrypted
exit
line ssh
 login authentication RUSATOM
 enable authentication RUSATOM 
 password da39a3ee5e6b4b0d3255bfef95601890afd80709 encrypted
exit
line console
 login authentication RUSATOM
 enable authentication RUSATOM
 password da39a3ee5e6b4b0d3255bfef95601890afd80709 encrypted
exit
!
enable password level 15 encrypted 01016d55a61409c7a29e4bd274ff510d7a66860b
!
username admin password encrypted ca0c3047353c12e25f9cb57d51e63fed69979740 privilege 15
!
ip ssh server
!
aaa authentication login RUSATOM local
aaa authentication enable RUSATOM enable
!
no ip telnet server
!
!
{conf_int}
!
interface gigabitethernet1/0/25
 description LAG_to_801-SCADA-AGG1-ctrl_{u1}
 storm-control broadcast level 20 trap shutdown
 storm-control multicast level 20 trap shutdown
 channel-group 1 mode auto
 lldp optional-tlv port-desc sys-name
exit
!
interface gigabitethernet1/0/26
 description LAG_to_801-SCADA-AGG1-ctrl_{u2}
 storm-control broadcast level 20 trap shutdown
 storm-control multicast level 20 trap shutdown
 channel-group 1 mode auto
 lldp optional-tlv port-desc sys-name
exit
!
!
interface Port-channel1
 description to_801-SCADA-AGG1-ctrl
 switchport mode trunk
 switchport trunk allowed vlan 2021,{vlans}
exit
!
interface vlan 1
 shutdown
exit
!
interface vlan 2021
 name MGMT_SCADA_ANE
 ip address {ip_add} 255.255.255.128
exit
!
ip route 0.0.0.0 0.0.0.0 10.52.211.129
!
end

'''


config_lines = []
interface_num = 1
current_device = None
current_hostname = None
current_ip = None
vlan_set = set()

# Cловарь для хранения данных об uplink'ах
uplinks = {}


# Перебираем строки, в листе 'upliks' и формируем словарь для хранения данных об uplink'ах
for row in sheet_uplinks.iter_rows(min_row=2, values_only=True):
    device_name = row[0]
    uplinks[device_name] = {'Port1': row[6], 'Port2': row[7]}


# Перебираем строки, в листе KJ начиная с четвертой
for row in sheet_main.iter_rows(min_row=4, max_row=276, min_col=6, max_col=20):
    hostname_old, ip_old, interface_old, description, mode, vlan, status, role, _, _, _, _, _, hostname_new, ip_new = [cell.value for cell in row]

    # Пропускаем строки без hostname или в состоянии 'shutdown', а также в роли 'chain' или 'uplink'
    if not hostname_old or not interface_old:
        continue
    if status == 'shutdown':
        continue
    if role == 'chain' or role == 'uplink':
        continue

    # Сохраняем конфигурацию для текущего устройства
    if hostname_old != current_device and current_device is not None:
        save_configuration(out_dir, current_hostname, current_ip, config_lines, vlan_set, uplink1, uplink2)
        
        # Обналяем счетчики
        config_lines = []
        interface_num = 1
        vlan_set = set()

    # Сохраняем данные следующего устройства
    current_device = hostname_old
    current_hostname = hostname_new
    current_ip = ip_new

    # Записываем номера VLAN для trunk интерфейсов
    vlan_set.add(int(vlan))

    # Создем интерфейс и прибавляем счетчик
    interface_name = f'GigabitEthernet1/0/{interface_num}'
    interface_num += 1
    uplink1 = uplinks[current_hostname]['Port1']
    uplink2 = uplinks[current_hostname]['Port2']
    
    if description is not None:
        description = description.replace(' ', '_')
    # Добовляем конфигурацию интрефейса в список
    config_lines.append(generate_interface_config(interface_name, description, mode, vlan, role))

    # Записываем созданный интерфейс в таблицу
    sheet_main.cell(row=row[0].row, column=21, value=interface_name)
   
    print(config_lines)

# Проверяем, осталась ли несохраненная конфигурация для последнего устройства
if current_device is not None:
    save_configuration(out_dir, current_hostname, current_ip, config_lines, vlan_set, uplink1, uplink2)

excel_book.save(new_excel_path)