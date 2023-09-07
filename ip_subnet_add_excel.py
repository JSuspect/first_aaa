import openpyxl
from pprint import pprint
import ipaddress

target_book = openpyxl.load_workbook(filename='/mnt/c/Users/SBakhvalov/Documents/_dev/wsl_dir/SB_ip_plan.xlsx', data_only=True)

tsheet = target_book['Fast Track']
address_list = []    # Создаем пустой список
data_dict = {}    # Создаем пустой словарь
subnet_dict = {}  # Создаем словарь для следующих подсетей /30

# Пример как можно сделать список списков из полученных данных (не используется далее)
for row in tsheet.iter_rows(min_row=3, max_row=tsheet.max_row, min_col=10, max_col=10):
    for cell in row:
        ip_addr = [cell.value, cell]
        address_list.append(ip_addr)

# Проходимся построчно по Листу 'Fast Track', в 10 столбеце. Наполняем подсетями словарь data_dict
for row in tsheet.iter_rows(min_row=3, max_row=tsheet.max_row, min_col=10, max_col=10):
    for cell in row:
        key = cell.coordinate  # Используем координату ячейки в качестве ключа (например, 'J3', 'J4' и т.д.)
        value = cell.value.strip()  # Используем значение ячейки в качестве значения
        data_dict[key] = value  # Добавляем пару ключ-значение в словарь

print('#' * 120) # Разделитель

# Генерируем словарь subnet_dict
# subnet_dict должен содержать следующую подсеть /30 от подсети которая содержится в словаре data_dict
for key, value in data_dict.items():
    try:
        ip_network = ipaddress.IPv4Network(value, strict=True) # Проходимся по значениям словаря и каждую итерацию записываем подсеть в ip_network
        next_subnet = ip_network.network_address + 4  # Следующая подсеть /30
        subnet_dict[key] = str(next_subnet) + '/30' # Записываем в словарь subnet_dict ключи полученные из data_dict и следующую подсеть /30
    except ipaddress.AddressValueError:
        pass

print('#' * 120) # Разделитель

for key in list(subnet_dict.keys()):  # Пройдемся по копии ключей (чтобы избежать изменения словаря во время итерации)
    if key.startswith('J'):
        new_key = 'R' + key[1:]  # Заменяем 'J' на 'Q'
        subnet_dict[new_key] = subnet_dict.pop(key)  # Изменяем ключ в словаре


for key, value in subnet_dict.items():
    tsheet[key].value = value                               # Записываем значение из subnet_dict в соответствующую ячейку на листе
    network = ipaddress.IPv4Network(value, strict=True)     # Вычисляем сеть из значения
     # Вычисляем первый и второй IP-адреса в сети
    first_host_ip = str(network.network_address + 1)
    second_host_ip = str(network.network_address + 2)
    print(first_host_ip, second_host_ip)

    # Записываем first_host_ip в соседний столбец (столбец )
    tsheet[key.replace('R', 'S')].value = first_host_ip
    
    # Записываем second_host_ip в соседний столбец (столбец C)
    tsheet[key.replace('R', 'T')].value = second_host_ip

print('#' * 120) # Разделитель

# Сохраняем книгу
target_book.save('/mnt/c/Users/SBakhvalov/Documents/_dev/wsl_dir/python_proc_ip_plan_SB.xlsx')