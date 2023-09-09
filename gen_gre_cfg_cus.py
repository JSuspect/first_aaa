import openpyxl
from openpyxl import load_workbook
import argparse
from transliterate import translit

# Парсинг аргументов командной строки
parser = argparse.ArgumentParser(description="Generate GRE tunnel configuration for Cisco ASR1002 routers.")
parser.add_argument("-r1", action="store_true", help="Generate configuration for router r1.")
parser.add_argument("-r2", action="store_true", help="Generate configuration for router r2.")
args = parser.parse_args()

# Путь к файлу Excel с данными
excel_file_path = '/mnt/c/Users/SBakhvalov/Documents/_dev/wsl_dir/pyProc_IP_plan_SB.xlsx'

# Открываем файл Excel с аргументом data_only=True
workbook = load_workbook(excel_file_path, data_only=True)
sheet = workbook.active

# Определяем шаблон конфигурации GRE туннеля для Cisco ASR1002
gre_template = """
interface Tunnel{tunnel_number}
 description {description}
 bandwidth 10000
 ip vrf forwarding VPN
 ip address {tunnel_ip} {subnet_mask}
 {delay_command}
 keepalive 10 3
 tunnel source {tunnel_src}
 tunnel destination {destination_ip}
 tunnel vrf {tun_vrf}
"""

# Функция для генерации конфигурации GRE туннелей
def generate_gre_config(router):
    # Проходимся по строкам файла Excel, начиная с второй строки (первая строка - заголовки)
    for row in sheet.iter_rows(min_row=3, values_only=True):
        remote_site = row[0]  # Название удаленной площадки на русском
        remote_site_translit = translit(remote_site, 'ru', reversed=True).upper()  # Транслитерация на английский
        split_site_translit = remote_site_translit.split()
        # Проверяем количество слов
        if len(split_site_translit) > 1:
            # Если есть более одного слова, добавляем тире между ними
            modified_remote_site_translit = "-".join(split_site_translit)
        else:
            # Если только одно слово, оставляем без изменений
            modified_remote_site_translit = remote_site_translit
        tun_source = row[21 if router == "r1" else 22]  # IP адрес туннельного интерфейса для r1 или r2
        tun_ip = row[10 if router == "r1" else 18]  # IP адрес туннельного интерфейса для r1 или r2
        tunnel_ip_parts = tun_ip.split('.')
        second_octet = tunnel_ip_parts[2]
        if len(second_octet) == 1:
            sec_o = '00' + second_octet
        elif len(second_octet) == 2:
            sec_o = '0' + second_octet
        else:
            sec_o = second_octet

        tunnel_number = tunnel_ip_parts[-3] + sec_o + tunnel_ip_parts[-1]
        subnet_mask = row[12 if router == "r1" else 20]  # Маска подсети для r1 или r2
        tun_destination = row[6 if router == "r1" else 15]  # IP адрес назначения для r1 или r2
        if router == "r1":
            description = f"-> RT01-{modified_remote_site_translit}-BRD via ТТК"
            tunnel_vrf = 'VPN'
        else:
            description = f"-> RT01-{modified_remote_site_translit}-BRD via MEGATON"
            tunnel_vrf = 'EXT'
        
        delay_command = "no delay" if router == "r1" else "delay 150000"


        # Генерируем конфигурацию для выбранного роутера
        gre_config = gre_template.format(
            tunnel_number=tunnel_number,  # Номер туннеля соответствует названию удаленной площадки
            description=description,
            tunnel_src=tun_source,
            tunnel_ip=tun_ip,
            subnet_mask=subnet_mask,
            destination_ip=tun_destination,
            delay_command=delay_command,
            tun_vrf=tunnel_vrf
        )

        # Выводим сгенерированную конфигурацию в терминал
        print(gre_config)

# Проверяем, какие роутеры указаны в аргументах и генерируем конфигурацию для них
if args.r1:
    router = "r1"
    print("#" * 20, router, "#" * 20)
    generate_gre_config(router)

if args.r2:
    router = "r2"
    print("#" * 20, router, "#" * 20)
    generate_gre_config(router)