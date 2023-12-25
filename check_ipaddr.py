import pathlib
import openpyxl
import ipaddress

path1 = pathlib.Path.home() / 'slink/wsld/cppk/Билайн_адреса.xlsx'
path2 = pathlib.Path.home() / 'slink/wsld/cppk/ТТК_адреса.xlsx'

bee_book = openpyxl.load_workbook(path1, data_only=True)
ttk_book = openpyxl.load_workbook(path2, data_only=True)

sheet_bee = bee_book.active
sheet_ttk = ttk_book.active

list_false_bee = []

for row in sheet_bee.iter_rows(min_row=2, values_only=True):
    station = row[0]
    ip1 = row[1]
    ip2 = row[2]
    ip_sation = ipaddress.IPv4Address(ip1.strip())
    ip_isp = ipaddress.IPv4Address(ip2.strip())
    if ip_sation == (ip_isp - 1):
        print(f'Станция: {station:<15}    IP_ADDR_STATION: {ip_sation}    IP_ADDR_ISP: {ip_isp} SUBNET: OK')
    else:
        list_false_bee.append(station)
        print(f'Станция: {station:<15}    IP_ADDR_STATION: {ip_sation}    IP_ADDR_ISP: {ip_isp} SUBNET: FALSE!!!')
print(list_false_bee)


list_false_ttk = []

for row in sheet_ttk.iter_rows(min_row=2, values_only=True):
    station = row[0]
    ip1 = row[1]
    ip2 = row[2]
    ip_sation = ipaddress.IPv4Address(ip1.strip())
    ip_isp = ipaddress.IPv4Address(ip2.strip())
    if ip_sation == (ip_isp - 1):
        print(f'Станция: {station:<15}    IP_ADDR_STATION: {ip_sation}    IP_ADDR_ISP: {ip_isp} SUBNET: OK')
    else:
        list_false_ttk.append(station)
        print(f'Станция: {station:<15}    IP_ADDR_STATION: {ip_sation}    IP_ADDR_ISP: {ip_isp} SUBNET: FALSE!!!')
print(list_false_ttk)
