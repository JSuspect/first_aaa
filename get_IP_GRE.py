import openpyxl
import os
import re

# Создаем новую книгу для результатов
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Указываем путь к папке, где находятся файлы Excel
folder_path = 'путь_к_папке'  # Замените на фактический путь к вашей папке

# Получаем список файлов Excel в указанной папке
excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

# Создаем список для хранения данных
additional_data = []

# Итерируемся по каждому файлу
for excel_file in excel_files:
    # Полный путь к текущему файлу
    full_path = os.path.join(folder_path, excel_file)

    # Открываем текущий файл Excel
    current_workbook = openpyxl.load_workbook(full_path)
    current_sheet = current_workbook.active

    # Ищем ячейку с текстом "Адресация GRE Tun:"
    target_text = "Адресация GRE Tun:"
    found = False

    for row in current_sheet.iter_rows():
        for cell in row:
            if cell.value == target_text:
                found = True
                target_row = cell.row
                target_column = cell.column

    # Если ячейка найдена, извлекаем IP-подсеть /30
    if found:
        extracted_data = []
        for row in current_sheet.iter_rows(min_row=target_row, max_row=current_sheet.max_row, min_col=target_column + 1, max_col=target_column + 1):
            for cell in row:
                # Извлекаем IP-подсеть /30
                ip_subnet = cell.value

                # Извлекаем IP-адреса хостов с использованием регулярных выражений
                ip_addresses = re.findall(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b', ip_subnet)

                if len(ip_addresses) == 2:
                    first_host_ip, second_host_ip = ip_addresses
                    extracted_data.append((first_host_ip, second_host_ip))
        
        # Добавляем название файла без расширения и извлеченные данные в список
        additional_data.append((os.path.splitext(excel_file)[0], extracted_data))

    # Закрываем текущий файл Excel
    current_workbook.close()

# Добавляем данные из списка `additional_data` в файл `output.xlsx`
for data in additional_data:
    for ip_pair in data[1]:
        output_sheet.append([data[0], ip_pair[0], ip_pair[1]])

# Сохраняем результаты в новом файле
output_workbook.save('output.xlsx')
