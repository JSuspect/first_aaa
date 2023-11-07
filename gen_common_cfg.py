from openpyxl import load_workbook
from pathlib import Path

# Шаблон для содержимого файла конфигурации
CONFIG_TEMPLATE = """hostname {hostname}
interface Vlan10
 ip address {ip} 255.255.255.128
"""

def create_config_files(excel_path, config_dir_path, start_row, end_row, ip_col, reserve_col, hostname_col, template):
    # Загрузка книги Excel
    workbook = load_workbook(filename=excel_path)
    sheet = workbook['IP plan']

    # Проверка и создание каталога для конфигураций, если он не существует
    Path(config_dir_path).mkdir(parents=True, exist_ok=True)

    # Прохождение по строкам и столбцам для извлечения данных и создания файлов конфигурации
    for row in range(start_row, end_row + 1):

        reserve_status = sheet.cell(row, reserve_col).value
        # Проверяем, не является ли текущая строка резервной
        if reserve_status == 'Резерв':
            continue  # Пропускаем создание файла для этой строки

        ip = sheet.cell(row, ip_col).value
        hostname = sheet.cell(row, hostname_col).value

        # Формирование содержимого файла конфигурации
        config_content = template.format(hostname=hostname, ip=ip)

        # Формирование имени файла и создание файла
        filename = f"{ip}_{hostname}.txt"
        file_path = Path(config_dir_path) / filename
        with open(file_path, 'w') as config_file:
            config_file.write(config_content)
        
        if file_path.exists() == True:
            print(f"Файл конфигурации {filename} создан.\nПолный путь к файлу: {file_path}\n")
        else:
            print(f"Файл конфигурации {filename} не создан.")


# Параметры
excel_file_path = '/mnt/c/Users/SBakhvalov/Documents/_wsl_dir/rusatom/t1.xlsx'
config_directory_path = '/mnt/c/Users/SBakhvalov/Documents/_wsl_dir/rusatom/config/py_gen_common_cfg'
start_row_index = 62
end_row_index = 89
ip_column_index = 7
hostname_column_index = 8
reserve_col = 9

# Вызов функции
create_config_files(
    excel_path=excel_file_path,
    config_dir_path=config_directory_path,
    start_row=start_row_index,
    end_row=end_row_index,
    ip_col=ip_column_index,
    hostname_col=hostname_column_index,
    template=CONFIG_TEMPLATE,
    reserve_col=reserve_col
)
